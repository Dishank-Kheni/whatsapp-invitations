# code to manipulate image
from PIL import Image, ImageFont, ImageDraw

from PIL import Image
from PIL import ImageDraw
from langdetect import detect
import openpyxl
import os
import sys
import json
from fpdf import FPDF
import io

jsondata = json.loads(sys.argv[1])

cwd = os.getcwd()


def imgshouldmanipulate(imgno):
    loopno = 0
    for data in jsondata:
        if data['imageno'] == imgno:
            return loopno

        else:
            loopno = loopno+1

    return -1
    # print(json.dumps(data))


f = open(cwd+'\invitation\meta_data.json', 'r')
metadata = json.load(f)
imagename = metadata['imgbasename']
excelname = metadata['excelname']
imgcount = int(metadata['noofimage'])

file = cwd + '\invitation\excel_sheet\\' + excelname
wb_obj = openpyxl.load_workbook(file)
sheet_obj = wb_obj.active
# print(sheet_obj.max_row)
max_col = sheet_obj.max_column

# first loop : excel
for i in range(2, sheet_obj.max_row+1):
    name = sheet_obj.cell(row=i, column=2).value
    mobileNo = sheet_obj.cell(row=i, column=4).value
    lan = detect(name)
    # print('r')
    pdf = FPDF()
    for j in range(imgcount):
        imgmanipulateno = imgshouldmanipulate(j)

        if imgmanipulateno != -1:
            # print("img manipulate")
            img = Image.open(
                cwd+'\invitation\original_images\\'+imagename + str(j)+'.jpg')
            img_copy = img.copy()
            I1 = ImageDraw.Draw(img_copy)

            if(lan == "gu"):
                name = sheet_obj.cell(row=i, column=3).value
                myFont = ImageFont.truetype(
                    'Shree768.ttf', int(jsondata[imgmanipulateno]['size']), layout_engine=ImageFont.LAYOUT_RAQM)
                I1.text((int(jsondata[imgmanipulateno]['coordinate'][0]), int(jsondata[imgmanipulateno]['coordinate'][1])), name,
                        font=myFont, fill=(int(jsondata[imgmanipulateno]['color'][0]), int(jsondata[imgmanipulateno]['color'][1]), int(jsondata[imgmanipulateno]['color'][2])))
                # img_copy.show()
                p = cwd + "\invitation\\temp\invitation_img\\" + \
                    imagename+str(j)+'.jpg'
                img_copy.save(p)
                pdf.add_page()
                pdf.image(p, 0, 0, 210, 297)
                # p = cwd + '\public\\test\\test.jpg'
            else:
                name = sheet_obj.cell(row=i, column=2).value
                # mno = str(mobileNo)+".jpg"
                # p = Path(working_directory/'invitation', mno)
                myFont = ImageFont.truetype(
                    'Roboto-Bold.ttf', int(jsondata[imgmanipulateno]['size']))
                I1.text((int(jsondata[imgmanipulateno]['coordinate'][0]), int(jsondata[imgmanipulateno]['coordinate'][1])), name,
                        font=myFont, fill=(int(jsondata[imgmanipulateno]['color'][0]), int(jsondata[imgmanipulateno]['color'][1]), int(jsondata[imgmanipulateno]['color'][2])))
                # img_copy.show()
                p = cwd + "\invitation\\temp\invitation_img\\" + \
                    imagename+str(j)+'.jpg'
                img_copy.save(p)
                pdf.add_page()
                pdf.image(p, 0, 0, 210, 297)

        else:
            # print("add dirctly to pdf")
            img = Image.open(
                cwd+'\invitation\original_images\\'+imagename + str(j)+'.jpg')
            img_copy = img.copy()
            p = cwd + "\invitation\\temp\invitation_img\\" + \
                imagename+str(j)+'.jpg'
            img_copy.save(p)
            pdf.add_page()
            pdf.image(p, 0, 0, 210, 297)
            # for data range()
    pdf.output(cwd + "\invitation\pdfs\\" + str(mobileNo)+".pdf", "F")
