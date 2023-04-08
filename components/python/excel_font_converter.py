import sys

from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import time
from bs4 import BeautifulSoup
from selenium.webdriver.support.ui import Select
import os
import openpyxl
from pathlib import Path
import json

fontURL = "https://www.pramukhfontconverter.com/gujarati"
shouldRun = True


def pramukhWebAutomation(Excel=2):
    try:
        driver = webdriver.Chrome(service=Service(
            ChromeDriverManager().install()))

        def remain_attemp():
            attempt_warning = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div/article/div[2]/div[2]")))
            attempts = attempt_warning.get_attribute("outerHTML")
            # print(attempts)
            soup = BeautifulSoup(attempts, 'lxml')
            attempt_text = soup.find('div').text
            # print(attempt_text)
            stList = attempt_text.split(" ")
            # print(  stList[9])
            return stList[9]

        def convertFont(name):

            # driver.get('https://www.pramukhfontconverter.com/gujarati')
            driver.execute_script("window.scrollTo(0, 250)")
            # /html/body/div[2]/div/article/div[2]/div[3]/div[1]/div/div[1]/select/option[73]
            # /html/body/div[2]/div/article/div[2]/div[3]/div[1]/div/div[1]/select
            # /html/body/div[2]/div/article/div[2]/div[3]/div[1]/div/div[2]/textarea
            # ddelement= Select(driver.find_element(By.XPATH, "/html/body/div[1]/div/article/div[2]/div[3]/div[3]/div/div[1]/select"))

            ddelement = Select(driver.find_element(
                By.XPATH, "/html/body/div[2]/div/article/div[2]/div[3]/div[3]/div/div[1]/select"))
            ddelement.select_by_visible_text('Shree Guj 768')
            select_text_box = WebDriverWait(driver, 10).until(
                # EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div/article/div[2]/div[3]/div[1]/div/div[2]/textarea")))

                EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div/article/div[2]/div[3]/div[1]/div/div[2]/textarea")))
            # names = ['પાર્થ','એલવીશ', 'રામાણી' ,  'યશ', 'રૂપની', 'પ્રીત']
            # for name in names:
            select_text_box.clear()
            select_text_box.send_keys(name)
            convert = WebDriverWait(driver, 10).until(
                # EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div/article/div[2]/div[3]/div[2]/button[1]")))
                EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div/article/div[2]/div[3]/div[2]/button[1]")))
            convert.click()
            time.sleep(5)
            result = driver.find_element_by_id(
                "txtUnicode").get_property('value')
            print(result)
            return result

        # /html/body/div[2]/div/article/div[2]/div[3]/div[2]/button[1]
        driver.get("https://www.hidemyass-freeproxy.com/")
        site_access_box = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/form/div[1]/div/input[2]")))
        # EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div/article/div[2]/div[3]/div[2]/button[1]")))
        site_access_box.click()
        site_access_box.clear()
        site_access_box.send_keys(fontURL)
        connectBtn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/form/div[3]/a")))
        connectBtn.click()

        attempt_number = int(remain_attemp())
        proxyChangeList = ["/html/body/div[1]/form/div[4]/div[2]/div[3]",
                           "/html/body/div[1]/form/div[4]/div[2]/div[4]",
                           "/html/body/div[1]/form/div[4]/div[2]/div[5]",
                           "/html/body/div[1]/form/div[4]/div[2]/div[6]",
                           "/html/body/div[1]/form/div[4]/div[2]/div[7]",
                           "/html/body/div[1]/form/div[4]/div[2]/div[8]", ]

        counter = 0
        ExcelContent = 40
        reachLevel = 2
        increase = 0
        # print(attempt_number)

        cwd = os.getcwd()
        f = open(cwd+'\invitation\meta_data.json', 'r')
        metadata = json.load(f)
        excelname = metadata['excelname']

        file = cwd + '\invitation\excel_sheet\\' + excelname
        # print(file)
        wb_obj = openpyxl.load_workbook(file)
        sheet_obj = wb_obj.active
        rows = sheet_obj
        print(sheet_obj.max_row)
        # max_col = sheet_obj.max_column
        for i in range(Excel, sheet_obj.max_row+1):
            # print(sheet_obj.cell(row=i,column=2).value)
            # while(reachLevel!=rows):
            time.sleep(2)
            attempt_number = int(remain_attemp())
            if(attempt_number <= 3):
                print("Change Proxy")
                if(counter >= 6):
                    counter = 0
                else:
                    counter = counter+1

                changeProxyBtn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/form/div[4]/div[1]")))
                changeProxyBtn.click()

                changeProxyBtn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, proxyChangeList[counter])))
                changeProxyBtn.click()

                agreeBtn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/a[2]")))
                agreeBtn.click()

            else:
                # print("Convert Font")
                # print("attempt number ")
                # print(attempt_number)
                # print("Convert text " + sheet_obj.cell(row=i, column=2).value)
                # print(" i value " + str(i))
                result = convertFont(sheet_obj.cell(row=i, column=2).value)
                time.sleep(1)
                # print("Convert font ")
                # print(reachLevel)

                c = sheet_obj.cell(row=i, column=3)
                c.value = result
                reachLevel = reachLevel+1
        wb_obj.save(file)
        return False
    except Exception as e:
        print("Exception")
        # print(reachLevel)
        print(str(e))
        driver.quit()
        return True
        pramukhWebAutomation(reachLevel)


# while(shouldRun):
pramukhWebAutomation(2)
print("Convertion successfully...")

sys.stdout.flush()
